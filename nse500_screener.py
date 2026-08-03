"""
NSE 500 Fundamental + Technical Screener - Core (v5, weekly cache + watchlist)
=====================================================================
Technicals for the shortlisted stocks are fetched in ONE batched call
via yf.download() instead of looping one Ticker() per stock - ~2
requests total instead of ~150-240.

Results are cached per ISO week (Mon-Sun). Running again within the
same week loads the cached result instantly with zero network calls;
a new week (or an explicit force-refresh from the dashboard) triggers
a fresh fetch.

v5 adds a lightweight per-symbol watchlist technicals fetch (LTP,
weekly EMA9/EMA11 of the High/Low series, weekly RSI) used to classify
each watchlist symbol as Fear / Extreme Fear / Neutral / Greed /
Extreme Greed. This is fetched and cached independently of the main
NSE500 screen (see db.watchlist_cache - shared across users, at most
once per calendar day per symbol).

REQUIREMENTS (run locally, needs internet):
    pip install yfinance pandas numpy openpyxl requests

USAGE:
    python nse500_screener.py
"""

import time
import io
import pickle
import sys
import threading
from datetime import date
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import numpy as np
import pandas as pd
import requests
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------------
OUTPUT_FILE = "NSE500_Screener.xlsx"
CACHE_DIR = Path(__file__).resolve().parent / ".cache"
CACHE_FILE = CACHE_DIR / "weekly_result.pkl"
MAX_WORKERS = 16
MIN_SECONDS_BETWEEN_CALLS = 0.5
MAX_RETRIES = 3
RETRY_BACKOFF_BASE = 5

# Breakout / retest signal configuration
BREAKOUT_RESISTANCE_LOOKBACK = 20
BREAKOUT_VOLUME_LOOKBACK = 20
BREAKOUT_VOLUME_MULTIPLIER = 1.5
BREAKOUT_RETEST_LOOKBACK = 5
BREAKOUT_RETEST_TOLERANCE = 0.01
BREAKOUT_MIN_HISTORY = BREAKOUT_RESISTANCE_LOOKBACK + 1

MIN_MCAP_CR = 10_000
MIN_ROCE = 15
MIN_SALES_GROWTH = 10
MIN_PROFIT_GROWTH = 10
MAX_DEBT_TO_EQUITY = 0.5

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
NSE500_CSV_URL = "https://archives.nseindia.com/content/indices/ind_nifty500list.csv"

print_lock = threading.Lock()
_rate_lock = threading.Lock()
_last_call_time = [0.0]


class CachedResult(dict):
    """Dictionary wrapper that keeps cache metadata while comparing like the payload only."""

    def __eq__(self, other):
        if isinstance(other, dict):
            other_payload = dict(other)
            self_payload = dict(self)
            for key in ("week", "saved_at"):
                self_payload.pop(key, None)
                other_payload.pop(key, None)
            return self_payload == other_payload
        return super().__eq__(other)


def _week_key(d=None):
    d = d or date.today()
    iso = d.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def ensure_cache_dir():
    CACHE_DIR.mkdir(exist_ok=True)


def save_daily_result(payload):
    ensure_cache_dir()
    payload = dict(payload)
    payload["week"] = _week_key()
    payload["saved_at"] = date.today().isoformat()
    with open(CACHE_FILE, "wb") as f:
        pickle.dump(CachedResult(payload), f)


def load_daily_result():
    if not CACHE_FILE.exists():
        return None
    try:
        with open(CACHE_FILE, "rb") as f:
            obj = pickle.load(f)
            if isinstance(obj, CachedResult):
                return obj
            if isinstance(obj, dict):
                return CachedResult(obj)
            return obj
    except Exception:
        return None


def should_refresh_cached_result(cached):
    if not cached:
        return True
    cached_week = cached.get("week")
    if not cached_week:
        return True
    try:
        return cached_week != _week_key()
    except Exception:
        return True


def rate_limit():
    with _rate_lock:
        wait = MIN_SECONDS_BETWEEN_CALLS - (time.time() - _last_call_time[0])
        if wait > 0:
            time.sleep(wait)
        _last_call_time[0] = time.time()


def with_retry(fn, *args, **kwargs):
    last_exc = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            last_exc = e
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_BACKOFF_BASE * attempt)
    raise last_exc


def get_nse500_symbols():
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36",
        "Accept": "text/csv,application/csv",
    }
    try:
        session = requests.Session()
        session.get("https://www.nseindia.com", headers=headers, timeout=10)
        resp = session.get(NSE500_CSV_URL, headers=headers, timeout=10)
        resp.raise_for_status()
        df = pd.read_csv(io.StringIO(resp.text))
        symbols = df["Symbol"].astype(str).str.strip().tolist()
        names = df["Company Name"].astype(str).str.strip().tolist()
        print(f"Fetched {len(symbols)} NSE 500 symbols from NSE.")
        return list(zip(symbols, names))
    except Exception as e:
        print(f"!! Could not auto-fetch NSE 500 list from NSE ({e}). "
              f"Falling back to local 'nse500_list.csv' if present.")
        try:
            df = pd.read_csv("nse500_list.csv")
            return list(zip(df["Symbol"].astype(str).str.strip(),
                             df["Company Name"].astype(str).str.strip()))
        except Exception as e2:
            print(f"!! No local fallback found either ({e2}). "
                  f"Create nse500_list.csv with columns Symbol,Company Name.")
            sys.exit(1)


def safe_get(d, key, default=np.nan):
    v = d.get(key, default) if hasattr(d, "get") else default
    return v if v is not None else default


def pct_growth(series):
    series = series.dropna()
    if len(series) < 2:
        return np.nan
    latest, prev = series.iloc[0], series.iloc[1]
    if prev == 0 or pd.isna(prev) or pd.isna(latest):
        return np.nan
    return round(((latest - prev) / abs(prev)) * 100, 2)


def _fetch_ticker_bundle(ticker_yf):
    rate_limit()
    t = yf.Ticker(ticker_yf)
    info = t.info or {}
    if not info:
        raise ValueError("empty info response (likely rate-limited)")
    fin = t.financials
    bs = t.balance_sheet
    return info, fin, bs


def get_fundamentals(symbol):
    ticker_yf = symbol + ".NS"
    row = {
        "Symbol": symbol, "NSE Ticker": ticker_yf,
        "Market Cap (Cr)": np.nan, "ROCE (%)": np.nan,
        "Sales Growth (%)": np.nan, "Profit Growth (%)": np.nan,
        "Debt to Equity": np.nan,
    }
    try:
        info, fin, bs = with_retry(_fetch_ticker_bundle, ticker_yf)

        mcap = safe_get(info, "marketCap")
        if pd.notna(mcap):
            row["Market Cap (Cr)"] = round(mcap / 1e7, 2)

        if fin is not None and not fin.empty and "Total Revenue" in fin.index:
            row["Sales Growth (%)"] = pct_growth(fin.loc["Total Revenue"])
        if fin is not None and not fin.empty and "Net Income" in fin.index:
            row["Profit Growth (%)"] = pct_growth(fin.loc["Net Income"])

        try:
            ebit = fin.loc["EBIT"].iloc[0] if "EBIT" in fin.index else np.nan
            total_assets = bs.loc["Total Assets"].iloc[0] if "Total Assets" in bs.index else np.nan
            current_liab = bs.loc["Current Liabilities"].iloc[0] if "Current Liabilities" in bs.index else np.nan
            if pd.notna(ebit) and pd.notna(total_assets) and pd.notna(current_liab):
                cap_employed = total_assets - current_liab
                if cap_employed != 0:
                    row["ROCE (%)"] = round((ebit / cap_employed) * 100, 2)
        except Exception:
            pass

        dte = safe_get(info, "debtToEquity")
        if pd.notna(dte):
            row["Debt to Equity"] = round(dte / 100, 2)
        else:
            try:
                total_debt = bs.loc["Total Debt"].iloc[0] if "Total Debt" in bs.index else np.nan
                equity = bs.loc["Common Stock Equity"].iloc[0] if "Common Stock Equity" in bs.index else np.nan
                if pd.notna(total_debt) and pd.notna(equity) and equity != 0:
                    row["Debt to Equity"] = round(total_debt / equity, 2)
            except Exception:
                pass

    except Exception as e:
        row["_error"] = str(e)
        with print_lock:
            print(f"   [warn] {symbol}: {e}")

    return row


def ema(series, span):
    return series.ewm(span=span, adjust=False).mean()


def rsi(series, period=14):
    delta = series.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1 / period, min_periods=period).mean()
    avg_loss = loss.ewm(alpha=1 / period, min_periods=period).mean()
    return 100 - (100 / (1 + avg_gain / avg_loss))


def supertrend(df, period=10, multiplier=1):
    hl2 = (df["High"] + df["Low"]) / 2
    tr = pd.concat([
        df["High"] - df["Low"],
        (df["High"] - df["Close"].shift()).abs(),
        (df["Low"] - df["Close"].shift()).abs(),
    ], axis=1).max(axis=1)
    atr = tr.ewm(alpha=1 / period, min_periods=period).mean()
    upperband = hl2 + multiplier * atr
    lowerband = hl2 - multiplier * atr
    st = pd.Series(index=df.index, dtype=float)
    direction = pd.Series(index=df.index, dtype=int)

    for i in range(len(df)):
        if i == 0:
            st.iloc[i] = upperband.iloc[i]
            direction.iloc[i] = 1
            continue
        if df["Close"].iloc[i] > upperband.iloc[i - 1]:
            direction.iloc[i] = 1
        elif df["Close"].iloc[i] < lowerband.iloc[i - 1]:
            direction.iloc[i] = -1
        else:
            direction.iloc[i] = direction.iloc[i - 1]
            if direction.iloc[i] == 1 and lowerband.iloc[i] < lowerband.iloc[i - 1]:
                lowerband.iloc[i] = lowerband.iloc[i - 1]
            if direction.iloc[i] == -1 and upperband.iloc[i] > upperband.iloc[i - 1]:
                upperband.iloc[i] = upperband.iloc[i - 1]
        st.iloc[i] = lowerband.iloc[i] if direction.iloc[i] == 1 else upperband.iloc[i]
    return st


def classify_breakout_retest_signal(df):
    """Return a signal label and a 0-100 score for a single symbol's OHLCV history.

    The function uses only historical data up to the latest row. It does not look
    at future candles and returns ``NONE`` whenever the required history is missing.
    """
    if df is None or df.empty:
        return "NONE", 0

    required = {"Open", "High", "Low", "Close", "Volume"}
    if not required.issubset(df.columns):
        return "NONE", 0

    hist = df.copy().sort_index()
    hist = hist.dropna(subset=["Open", "High", "Low", "Close", "Volume"]).reset_index(drop=True)
    if len(hist) < BREAKOUT_MIN_HISTORY:
        return "NONE", 0

    latest = hist.iloc[-1]
    prev = hist.iloc[-BREAKOUT_RESISTANCE_LOOKBACK - 1:-1] if len(hist) > BREAKOUT_RESISTANCE_LOOKBACK else None
    if prev is None or prev.empty:
        return "NONE", 0

    prev_highs = prev["High"]
    resistance = prev_highs.max()
    avg_volume = prev["Volume"].mean()
    ema50 = ema(hist["Close"], 50).iloc[-1]
    ema200 = ema(hist["Close"], 200).iloc[-1]
    rsi14 = rsi(hist["Close"], 14).iloc[-1]

    if pd.isna(ema50) or pd.isna(ema200) or pd.isna(rsi14):
        return "NONE", 0

    breakout_cond = (
        latest["Close"] > resistance and
        latest["Volume"] > BREAKOUT_VOLUME_MULTIPLIER * avg_volume and
        latest["Close"] > ema50 and
        ema50 > ema200
    )

    breakout_window = hist.iloc[-BREAKOUT_RETEST_LOOKBACK - 1:-1] if len(hist) > BREAKOUT_RETEST_LOOKBACK else None
    retest_cond = False
    if breakout_window is not None and not breakout_window.empty:
        breakout_level = breakout_window["Close"].max()
        retest_cond = (
            latest["Low"] >= breakout_level * (1 - BREAKOUT_RETEST_TOLERANCE) and
            latest["Low"] <= breakout_level * (1 + BREAKOUT_RETEST_TOLERANCE) and
            latest["Close"] > latest["Open"] and
            latest["Close"] > ema50 and
            ema50 > ema200
        )

    volume_increasing = hist["Volume"].iloc[-3:].mean() > hist["Volume"].iloc[-10:-3].mean() if len(hist) >= 10 else False
    breakout_ready_cond = (
        latest["Close"] >= resistance * 0.98 and
        latest["Close"] < resistance and
        volume_increasing and
        rsi14 > 55 and
        ema50 > ema200
    )

    if breakout_cond:
        label = "BREAKOUT"
    elif retest_cond:
        label = "RETEST"
    elif breakout_ready_cond:
        label = "BREAKOUT_READY"
    else:
        label = "NONE"

    score = 0
    if latest["Close"] > ema50:
        score += 15
    if ema50 > ema200:
        score += 15
    if rsi14 > 55:
        score += 10
    if latest["Volume"] > BREAKOUT_VOLUME_MULTIPLIER * avg_volume:
        score += 15
    if latest["Close"] > resistance:
        score += 25
    if retest_cond:
        score += 20

    score = max(0, min(100, score))
    return label, score


def _extract_symbol_frame(batch_df, ticker_yf, single_ticker):
    if single_ticker:
        return batch_df
    if ticker_yf in batch_df.columns.get_level_values(0):
        return batch_df[ticker_yf]
    return pd.DataFrame()


def get_technicals_batch(symbols, suffix=".NS"):
    """Fetch LTP + RSI/Supertrend/EMA (monthly & weekly) for ALL given
    symbols in just 2 network calls total. `suffix` is appended to each
    symbol to build the yfinance ticker - use '.NS' for NSE stocks
    (default) or '' for tickers that are already complete, like crypto
    pairs (e.g. 'BTC-USD')."""
    tickers = [s + suffix for s in symbols]
    single = len(tickers) == 1
    results = {s: {
        "LTP": np.nan, "RSI (Monthly)": np.nan, "Supertrend (Monthly)": np.nan,
        "EMA 20 (Monthly)": np.nan, "EMA 10 (Monthly)": np.nan,
        "RSI (Weekly)": np.nan, "Supertrend (Weekly)": np.nan,
        "EMA 20 (Weekly)": np.nan, "EMA 10 (Weekly)": np.nan,
    } for s in symbols}

    try:
        monthly = with_retry(
            yf.download, tickers=tickers, period="5y", interval="1mo",
            group_by="ticker", threads=True, progress=False, auto_adjust=True,
        )
    except Exception as e:
        print(f"   [warn-tech] monthly batch download failed: {e}")
        monthly = None

    time.sleep(1.5)

    try:
        weekly = with_retry(
            yf.download, tickers=tickers, period="2y", interval="1wk",
            group_by="ticker", threads=True, progress=False, auto_adjust=True,
        )
    except Exception as e:
        print(f"   [warn-tech] weekly batch download failed: {e}")
        weekly = None

    for sym in symbols:
        ticker_yf = sym + suffix
        try:
            hist_m = _extract_symbol_frame(monthly, ticker_yf, single).dropna() if monthly is not None else pd.DataFrame()
            hist_w = _extract_symbol_frame(weekly, ticker_yf, single).dropna() if weekly is not None else pd.DataFrame()
            if hist_m.empty or hist_w.empty:
                continue

            out = results[sym]
            out["LTP"] = round(float(hist_w["Close"].iloc[-1]), 2)

            out["RSI (Monthly)"] = round(rsi(hist_m["Close"]).iloc[-1], 2)
            out["Supertrend (Monthly)"] = round(supertrend(hist_m, 10, 1).iloc[-1], 2)
            out["EMA 20 (Monthly)"] = round(ema(hist_m["Close"], 20).iloc[-1], 2)
            out["EMA 10 (Monthly)"] = round(ema(hist_m["Close"], 10).iloc[-1], 2)

            out["RSI (Weekly)"] = round(rsi(hist_w["Close"]).iloc[-1], 2)
            out["Supertrend (Weekly)"] = round(supertrend(hist_w, 10, 1).iloc[-1], 2)
            out["EMA 20 (Weekly)"] = round(ema(hist_w["Close"], 20).iloc[-1], 2)
            out["EMA 10 (Weekly)"] = round(ema(hist_w["Close"], 10).iloc[-1], 2)

            signal, score = classify_breakout_retest_signal(hist_w)
            out["Breakout_Retest_Signal"] = signal
            out["Breakout_Score"] = score
        except Exception as e:
            with print_lock:
                print(f"   [warn-tech] {sym}: {e}")

    return results


def classify_market_condition(ltp, ema9_low, ema11_low, ema9_high, ema11_high, rsi_weekly):
    """Fear: LTP below both the weekly EMA9-of-Low and EMA11-of-Low.
    Greed: LTP above both the weekly EMA9-of-High and EMA11-of-High.
    Extreme variants additionally require weekly RSI confirmation
    (<30 for extreme fear, >70 for extreme greed). Anything else is
    Neutral."""
    if pd.isna(ltp):
        return "N/A"

    is_fear = pd.notna(ema9_low) and pd.notna(ema11_low) and ltp < ema9_low and ltp < ema11_low
    is_greed = pd.notna(ema9_high) and pd.notna(ema11_high) and ltp > ema9_high and ltp > ema11_high

    if is_fear:
        return "Extreme Fear" if pd.notna(rsi_weekly) and rsi_weekly < 30 else "Fear"
    if is_greed:
        return "Extreme Greed" if pd.notna(rsi_weekly) and rsi_weekly > 70 else "Greed"
    return "Neutral"


def _download_ohlcv_history(symbol, suffix=".NS", period="2y", interval="1wk", download_fn=None):
    """Fetch OHLCV history for a symbol, trying a fallback ticker format if needed.

    Some ETF-style tickers in the NSE watchlist (for example SETFNIF50 and
    SILVERBEES) may return empty history when requested as `SYMBOL.NS` but
    succeed when requested without the `.NS` suffix. This helper preserves the
    normal NSE path first and retries with the bare symbol as a fallback.
    """
    if download_fn is None:
        download_fn = yf.download

    candidates = [symbol + suffix]
    if suffix and not symbol.endswith(suffix):
        candidates.append(symbol)

    last_error = None
    for ticker in candidates:
        try:
            rate_limit()
            frame = with_retry(
                download_fn,
                tickers=ticker,
                period=period,
                interval=interval,
                progress=False,
                auto_adjust=True,
                group_by="column",
            )
            if isinstance(frame, pd.DataFrame) and not frame.empty:
                if isinstance(frame.columns, pd.MultiIndex):
                    if ticker in frame.columns.get_level_values(-1):
                        frame = frame.xs(ticker, axis=1, level=-1)
                    else:
                        frame.columns = frame.columns.get_level_values(0)
                return frame.dropna()
            last_error = ValueError(f"empty history returned for {ticker}")
        except Exception as exc:
            last_error = exc

    if last_error is not None:
        raise last_error
    raise ValueError(f"no history found for {symbol}")


def get_watchlist_technicals(symbol, suffix=".NS"):
    """Lightweight single-symbol fetch used for the Watchlist tab: just
    enough weekly history to compute LTP, weekly EMA9/EMA11 of the
    High/Low series, and weekly RSI, then classify Fear/Greed."""
    ticker_yf = symbol + suffix
    out = {
        "Symbol": symbol, "LTP": np.nan, "Condition": "N/A",
        "EMA 9 Low (Weekly)": np.nan, "EMA 11 Low (Weekly)": np.nan,
        "EMA 9 High (Weekly)": np.nan, "EMA 11 High (Weekly)": np.nan,
        "RSI (Weekly)": np.nan, "_error": None,
    }
    try:
        weekly = _download_ohlcv_history(symbol, suffix=suffix, period="2y", interval="1wk")
        if weekly.empty:
            raise ValueError(f"no weekly history returned for {ticker_yf}")

        ltp = round(float(weekly["Close"].iloc[-1]), 2)
        ema9_low = round(float(ema(weekly["Low"], 9).iloc[-1]), 2)
        ema11_low = round(float(ema(weekly["Low"], 11).iloc[-1]), 2)
        ema9_high = round(float(ema(weekly["High"], 9).iloc[-1]), 2)
        ema11_high = round(float(ema(weekly["High"], 11).iloc[-1]), 2)
        rsi_w = rsi(weekly["Close"]).iloc[-1]
        rsi_w = round(float(rsi_w), 2) if pd.notna(rsi_w) else np.nan

        out.update({
            "LTP": ltp,
            "EMA 9 Low (Weekly)": ema9_low, "EMA 11 Low (Weekly)": ema11_low,
            "EMA 9 High (Weekly)": ema9_high, "EMA 11 High (Weekly)": ema11_high,
            "RSI (Weekly)": rsi_w,
            "Condition": classify_market_condition(ltp, ema9_low, ema11_low, ema9_high, ema11_high, rsi_w),
        })
    except Exception as e:
        out["_error"] = str(e)
        with print_lock:
            print(f"   [warn-watchlist] {symbol}: {e}")
    return out


def get_watchlist_batch(symbols, suffix=".NS", max_workers=8):
    """Fetches get_watchlist_technicals() for each symbol in parallel.
    Returns {symbol: data_dict}. Intended to be called only for symbols
    that are missing from today's shared cache (see db.watchlist_cache)."""
    results = {}
    if not symbols:
        return results
    with ThreadPoolExecutor(max_workers=min(max_workers, len(symbols))) as pool:
        futures = {pool.submit(get_watchlist_technicals, s, suffix): s for s in symbols}
        for fut in as_completed(futures):
            r = fut.result()
            results[r["Symbol"]] = r
    return results


CRYPTO_TOP50 = [
    "BTC-USD", "ETH-USD", "USDT-USD", "BNB-USD", "SOL-USD", "XRP-USD", "USDC-USD",
    "ADA-USD", "AVAX-USD", "DOGE-USD", "TRX-USD", "DOT-USD", "MATIC-USD", "LTC-USD",
    "SHIB-USD", "LINK-USD", "BCH-USD", "NEAR-USD", "UNI-USD", "ATOM-USD", "XLM-USD",
    "ICP-USD", "ETC-USD", "FIL-USD", "HBAR-USD", "APT-USD", "ARB-USD", "VET-USD",
    "OP-USD", "MKR-USD", "INJ-USD", "IMX-USD", "GRT-USD", "RNDR-USD", "AAVE-USD",
    "ALGO-USD", "QNT-USD", "EGLD-USD", "SAND-USD", "MANA-USD", "THETA-USD", "AXS-USD",
    "XTZ-USD", "EOS-USD", "FLOW-USD", "CHZ-USD", "KAVA-USD", "XMR-USD", "CRV-USD",
    "LDO-USD", "FTM-USD",
]

CRYPTO_CACHE_FILE = CACHE_DIR / "weekly_crypto_result.pkl"


def save_crypto_result(df):
    ensure_cache_dir()
    payload = {"df": df, "week": _week_key(), "saved_at": date.today().isoformat()}
    with open(CRYPTO_CACHE_FILE, "wb") as f:
        pickle.dump(payload, f)


def load_crypto_result():
    if not CRYPTO_CACHE_FILE.exists():
        return None
    try:
        with open(CRYPTO_CACHE_FILE, "rb") as f:
            return pickle.load(f)
    except Exception:
        return None


def should_refresh_crypto_result(cached):
    if not cached:
        return True
    return cached.get("week") != _week_key()


def get_crypto_screener():
    """Fetches technicals for the top-50 crypto list and returns a
    dataframe sorted by Green Hits - same technical conditions as the
    NSE screener, no fundamentals stage (doesn't apply to crypto)."""
    symbols = [t.replace("-USD", "") for t in CRYPTO_TOP50]
    tech_results = get_technicals_batch(symbols, suffix="-USD")
    rows = [dict(tech, Symbol=sym) for sym, tech in tech_results.items()]
    df = pd.DataFrame(rows)

    hits = pd.DataFrame({name: df.apply(
        lambda r: bool(pd.notna(r.get("LTP")) and pd.notna(r.get(name)) and fn(r)), axis=1)
        for name, fn in CONDITION_MAP.items()})
    df["Green Hits"] = hits.sum(axis=1)
    df = df.sort_values("Green Hits", ascending=False).reset_index(drop=True)
    return df


def get_current_prices(symbols, suffix=".NS"):
    """Lightweight batched LTP-only lookup (no indicators) - used for
    paper-trade P&L, where all we need is the current price for a
    handful of symbols, not the full technicals fetch."""
    if not symbols:
        return {}
    tickers = [s + suffix for s in symbols]
    single = len(tickers) == 1
    try:
        data = with_retry(
            yf.download, tickers=tickers, period="5d", interval="1d",
            group_by="ticker", threads=True, progress=False, auto_adjust=True,
        )
    except Exception as e:
        print(f"   [warn] current price batch fetch failed: {e}")
        return {}

    prices = {}
    for sym in symbols:
        ticker_yf = sym + suffix
        try:
            frame = _extract_symbol_frame(data, ticker_yf, single).dropna()
            if not frame.empty:
                prices[sym] = round(float(frame["Close"].iloc[-1]), 2)
        except Exception:
            pass
    return prices


CONDITION_MAP = {
    "Supertrend (Monthly)": lambda r: r["LTP"] > r["Supertrend (Monthly)"],
    "EMA 20 (Monthly)": lambda r: r["LTP"] > r["EMA 20 (Monthly)"],
    "EMA 10 (Monthly)": lambda r: r["LTP"] > r["EMA 10 (Monthly)"],
    "RSI (Monthly)": lambda r: r["RSI (Monthly)"] < 30,
    "Supertrend (Weekly)": lambda r: r["LTP"] > r["Supertrend (Weekly)"],
    "EMA 20 (Weekly)": lambda r: r["LTP"] > r["EMA 20 (Weekly)"],
    "EMA 10 (Weekly)": lambda r: r["LTP"] > r["EMA 10 (Weekly)"],
    "RSI (Weekly)": lambda r: r["RSI (Weekly)"] < 30,
}


def write_excel(df1, df2, df3, output_file=OUTPUT_FILE):
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="All NSE500", index=False)
        df2.to_excel(writer, sheet_name="Fundamentally Sound", index=False)
        df3.to_excel(writer, sheet_name="Technical Screen", index=False)

    wb = load_workbook(output_file)
    ws = wb["Technical Screen"]
    headers = [c.value for c in ws[1]]
    for r_idx, row in df3.iterrows():
        excel_row = r_idx + 2
        for col_name, cond_fn in CONDITION_MAP.items():
            if col_name not in headers:
                continue
            col_idx = headers.index(col_name) + 1
            try:
                if pd.notna(row["LTP"]) and pd.notna(row.get(col_name)) and cond_fn(row):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_FONT
            except Exception:
                pass

    for ws_ in wb.worksheets:
        for col_cells in ws_.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            col_letter = get_column_letter(col_cells[0].column)
            ws_.column_dimensions[col_letter].width = min(max(length + 2, 10), 30)
    wb.save(output_file)


def main():
    cached = load_daily_result()
    if cached is not None and not should_refresh_cached_result(cached):
        df1, df2, df3 = cached["df_all"], cached["df_sound"], cached["df_tech"]
        print(f"Using this week's cached result (week {cached.get('week')}, "
              f"saved {cached.get('saved_at')}). Delete .cache/weekly_result.pkl "
              f"to force a refresh.")
        write_excel(df1, df2, df3 if df3 is not None else df2)
        print(f"Done. Saved to {OUTPUT_FILE}")
        return

    symbols = get_nse500_symbols()
    name_map = {sym: name for sym, name in symbols}

    print(f"\nFetching fundamentals for {len(symbols)} stocks ({MAX_WORKERS} workers)...")
    rows = []
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(get_fundamentals, s): s for s, _ in symbols}
        for i, fut in enumerate(as_completed(futures), 1):
            sym = futures[fut]
            r = fut.result()
            r["Company Name"] = name_map.get(sym, "")
            rows.append(r)
            if i % 25 == 0 or i == len(symbols):
                print(f"  {i}/{len(symbols)} done")
    print(f"Fundamentals done in {time.time()-t0:.0f}s")

    df1 = pd.DataFrame(rows)
    df1 = df1[["Symbol", "Company Name", "NSE Ticker", "Market Cap (Cr)", "ROCE (%)",
               "Sales Growth (%)", "Profit Growth (%)", "Debt to Equity"]]

    df2 = df1[
        (df1["Market Cap (Cr)"] > MIN_MCAP_CR) & (df1["ROCE (%)"] > MIN_ROCE) &
        (df1["Sales Growth (%)"] > MIN_SALES_GROWTH) & (df1["Profit Growth (%)"] > MIN_PROFIT_GROWTH) &
        (df1["Debt to Equity"] < MAX_DEBT_TO_EQUITY)
    ].copy().reset_index(drop=True)
    print(f"\n{len(df2)} stocks passed fundamentals.")

    if df2.empty:
        empty_tech = df2.assign(**{c: [] for c in CONDITION_MAP})
        write_excel(df1, df2, empty_tech)
        save_daily_result({"df_all": df1, "df_sound": df2, "df_tech": empty_tech})
        print("No stocks passed - nothing to run technicals on.")
        return

    print(f"\nFetching technicals for {len(df2)} stocks (batched)...")
    t1 = time.time()
    tech_results = get_technicals_batch(list(df2["Symbol"]))
    print(f"Technicals done in {time.time()-t1:.0f}s")

    tech_rows = [dict(tech, Symbol=sym) for sym, tech in tech_results.items()]
    df3 = df2.merge(pd.DataFrame(tech_rows), on="Symbol", how="left")

    if "Breakout_Retest_Signal" not in df3.columns:
        df3["Breakout_Retest_Signal"] = "NONE"
    if "Breakout_Score" not in df3.columns:
        df3["Breakout_Score"] = 0
    df3["Breakout_Retest_Signal"] = df3["Breakout_Retest_Signal"].fillna("NONE")
    df3["Breakout_Score"] = pd.to_numeric(df3["Breakout_Score"], errors="coerce").fillna(0)

    hits = pd.DataFrame({name: df3.apply(
        lambda r: bool(pd.notna(r.get("LTP")) and pd.notna(r.get(name)) and fn(r)), axis=1)
        for name, fn in CONDITION_MAP.items()})
    df3["Green Hits"] = hits.sum(axis=1)
    df3 = df3.sort_values(["Breakout_Score", "Green Hits"], ascending=[False, False]).reset_index(drop=True)

    write_excel(df1, df2, df3)
    save_daily_result({"df_all": df1, "df_sound": df2, "df_tech": df3})
    print(f"\nDone. Saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()